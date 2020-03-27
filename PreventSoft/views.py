from django.shortcuts import render, redirect
from django.forms import inlineformset_factory
from django.views.generic.list import ListView
from django.http import HttpResponse
from datetime import datetime

from django.contrib import admin, messages
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.models import User

from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.styles import GradientFill, Alignment, Color, colors
from openpyxl.workbook import Workbook

from tool.models import Tool
from EPP.models import EPP
from hazard.models import Hazard
from precaution.models import Precaution
from APR.models import APR, APRLine, set_documentnumber

from .forms import AprForm, AprLineForm, RegisterForm, EditUserForm
from .forms import AprEditForm, EppForm, HazardForm, PrecautionForm
from .forms import ToolForm, EppEditForm, ToolEditForm, AprLineEditForm


#### Index View.  
def index(request):
	#Checks if the user is already logged and redirects
	#to the document view
	if request.user.is_authenticated:
		return redirect('document')
	#Get data from login form and check with DB
	if request.method == 'POST':
		username = request.POST.get('username')
		password = request.POST.get('password')
		user = authenticate(username=username, password=password)

		#If the user exists and the data was correct 
		#then we proceed to login
		if user:
			login(request,user)
			messages.success(request, 'Bienvenido {}'.format(user.username))
			return redirect('document')
		#if an error occurred we redirect it to the 
		#same view and warn them.	
		else:
			messages.error(request, 'Usuario o contraseña no validos')	
	  
	return render(request, 'index.html', {'title' : 'Login',})

#### End of Index View

#### Views from User class
#This view is to create a new user. Only admin users
#can access it
def user_register(request):
	#We define the form.
	form = RegisterForm()
	if request.method == 'POST':
		form = RegisterForm(request.POST)
		#Check data
		if form.is_valid():
			#Delay the commit to make the password hash correctly.
			user = form.save(commit=False)
			user.set_password(user.password)
			#If the user going to be an admin
			#we set is Staff too.
			if user.is_superuser:
				user.is_staff = True
			#We save the data	
			user.save()
			#We check if the user was created and notify to the user
			if user:
				messages.success(request, 'Usuario {}, creado correctamente'.format(user.username))
				return redirect('user_register')
			else:
				messages.error(request, 'Ocurrió un error y el usuario no fué creado.')
	
	return render(request, 'users/user_register.html', 
						{ 'form' : form, 'title' : 'Nuevo Usuario', })	

#This view is to edit a user. Only admin users
#can access it
def user_edit(request, pk):
	#We take from url the pk of the user and 
	#generate an instance.
	instance = User.objects.get(pk=pk)
	#We define the form.
	form = EditUserForm()
	if request.method == 'POST':
		form = EditUserForm(request.POST, instance=instance)
		#Check data.
		if form.has_changed():
			if form.is_valid():
				#Delay the commit to make the password hash correctly.
				user = form.save(commit=False)
				user.set_password(user.password)
				#If the user going to be a admin
				#we set is Staff too.
				if user.is_superuser:
					user.is_staff = True
				#We save the data	
				user.save()
				#We check if the user was edited and notify to the user
				if user:
					messages.success(request, 'Usuario {}, editado correctamente'.format(user.username))
					return redirect('user_list')
				else:
					messages.error(request, 'Ocurrió un error y el usuario no fué editado.')
			else:
				messages.error(request, 'Ocurrió un error el formulario no es válido.')
		else:
			messages.success(request, 'Ningún cambio registrado.')
			return redirect('user_list')		
	#we pass the form with the instance to the view.
	form = EditUserForm(instance=instance)		
	return render(request, 'users/user_edit.html', 
					{ 'form' : form, 'title' : 'Editar Usuario', })

#This view is to delete a user. Only admin users
#can access it
def user_delete(request, pk):
	#We check if the logged user is admin
	if request.user.is_authenticated and request.user.is_superuser:
		#We take from url the pk of the user
		user = User.objects.get(pk=pk)
		#We delete the user and check
		if user.delete():
			messages.success(request, 'Usuario {}, eliminado correctamente'.format(user.username))
			return redirect('user_list')
		else:
			messages.error(request, 'Ocurrió un error y el usuario no fué eliminado.')
	else:
		return redirect('document')		

#This view shows in a list all users to edit and delete them.
#Is a view based in class.
class UserListView(ListView):
	template_name = 'users/user_list.html'
	queryset = User.objects.all().order_by('-id')
	#We rewrite the context data to send to the view.
	def get_context_data(self, **kwargs):
		context = super().get_context_data(**kwargs)
		context['message'] = 'Listado de Usuarios'
		context['title'] = 'Listado de Usuarios'
		context['list'] = context['user_list']
		
		return context		

#This view is to log out a user who requests it.
def logout_view(request):
	logout(request)
	messages.success(request, "Sesión cerrada con éxito")
	return redirect('index')	

#### End of Views from User class

#### Views from APR and APRLine class	
#This view shows the new document forms.
#It´s the default view when you login
def document_view(request):
	#We convert the Apr Line Form to a Inline Formset
	#to be able to save several document lines from a single view.
	LineFormSet = inlineformset_factory(
		APR,
		APRLine,
		form=AprLineForm,
		fields=('activities','tools','epps','hazards','precautions'),
		extra=5,
		max_num=8)
	#We define the form for document header.
	form = AprForm()
	formset = LineFormSet
	if request.method == 'POST':
		form = AprForm(request.POST)
		#Check data.
		if form.is_valid():
			#Delay the commit to set the document number and
			#user.
			data = form.save(commit=False)
			if data:
				data.documentnumber = set_documentnumber()
				data.user = request.user
				#We save data and keep it to instantiate each document line
				#with the correct document header.
				data.save()
				if request.method == 'POST':
					formset = LineFormSet(request.POST, instance=data)
					#Check data.
					if formset.is_valid():
						#Check if data was saved and notify to user
						if formset.save():
							messages.success(request, 
								'El documento {} se guardó correctamente'.format(data.documentnumber))
							return redirect('document')
						else:
							messages.error(request, 'Ocurrió un error y el documento no fué guardado.')	

	return render(request,'documents/document.html', 
					{ 'form': form,'formset': formset, 'title' : 'Nuevo APR'})

#This view shows in a list all documents headers to edit them.
#Is a view based in class.	
class DocumentListView(ListView):
	template_name = 'documents/document_list.html'
	queryset = APR.objects.all().order_by('-id')
	#We rewrite the context data to send to the view.
	def get_context_data(self, **kwargs):
		context = super().get_context_data(**kwargs)
		context['message'] = 'Listado de Documentos'
		context['title'] = 'Listado de Documentos'
		context['aprs'] = context['apr_list']
		
		return context

#This view is to edit a document. Only admin users
#can access it
def document_edit(request, pk):
	#We take from url the pk of the document and 
	#generate an instance.
	document = APR.objects.get(pk=pk)
	#We convert the Apr Line Form to a Inline Formset
	#to be able to save several document lines from a single view.
	LineFormSet = inlineformset_factory(
		APR,
		APRLine,
		form=AprLineEditForm,
		fields=('activities','tools','epps','hazards','precautions'),
		extra=5,
		max_num=8)
	#We define the form for document header.
	form = AprEditForm()
	formset = LineFormSet
	error = False
	alert = ''
	if request.method == 'POST':
		form = AprEditForm(request.POST, instance=document)
		#Check data.
		if form.has_changed():
			if form.is_valid():
				data = form.save()
				#Check if header was saved, then continue with lines.
				if data:
					alert = alert + ' ¡Se guardaron los cambios en la cabecera del Documento! '
			else:
				alert = alert + ' ¡Hubo un error con la cabecera del documento! '
				error = True
		else:
			alert = alert + ' ¡La cabecera del Documento no cambió! '

		if request.method == 'POST':
			formset = LineFormSet(request.POST, instance=document)
			#Check data.
			if formset.has_changed():
				if formset.is_valid():
					#Check if data was saved and notify to user
					line = formset.save()
					if line:
						alert = alert + ' ¡Se guardaron los cambios en las líneas del Documento! '
				else:
					alert = alert + ' ¡Hubo un error con las líneas del documento! '
					error = True
			else:
				alert = alert + ' ¡Las líneas del Documento no cambiaron! '	

		if error:
			messages.error(request, alert)
		else:
			messages.success(request, alert)
			return redirect('document_list')
				
	#We pass instantiated forms to the view.
	form = AprEditForm(instance=document)
	formset = LineFormSet(instance=document)

	return render(request,'documents/document_edit.html', 
					{ 'form': form,'formset': formset, 'title' : 'Editar APR'})

#This view generate a static table with data
#of the selected document.  
def document_detail(request, pk):
	#We take from url the pk of the document and 
	#keep data to send to view.
	document = APR.objects.get(pk=pk)
	documentLine = APRLine.objects.filter(apr_id=pk).order_by('id')

	return render(request,'documents/document_detail.html', 
					{ 'document': document,'lines': documentLine, 
						'title' : 'Detalles del APR'})

#### End of Views from APR and APRLine class	

#### Views from EPP class	
#This view is to create a new epp. Only admin users
#can access it
def epp_new(request):
	#We define the form.
	form = EppForm()
	if request.method == 'POST':
		form = EppForm(request.POST, request.FILES)
		#Check data.
		if form.is_valid():
			#Save data
			epp = form.save()
			#We check if the epp was created and notify to the user.
			if epp:
				messages.success(request, 'Equipo de Protección Personal creado correctamente')
				return redirect('epp_new')
			else:
				messages.error(request, 'Ocurrió un error y el E.P.P. no fué creado.')
	
	return render(request, 'epps/epp_new.html', 
					{ 'form' : form, 'title' : 'Nuevo E.P.P.'})

#This view shows in a list all epp´s to edit and delete them.
#Is a view based in class.
class EppListView(ListView):
	template_name = 'epps/epp_list.html'
	queryset = EPP.objects.all().order_by('name')
	#We rewrite the context data to send to the view.
	def get_context_data(self, **kwargs):
		context = super().get_context_data(**kwargs)
		context['message'] = 'Listado de Equipos de Protección Personal'
		context['title'] = 'Listado de E.P.P.'
		context['epps'] = context['epp_list']
		
		return context

#This view is to edit a epp. Only admin users
#can access it
def epp_edit(request, pk):
	#We take from url the pk of the epp and 
	#generate an instance.
	instance = EPP.objects.get(pk=pk)
	#We define the form.
	form = EppEditForm()
	if request.method == 'POST':
		form = EppEditForm(request.POST, request.FILES, instance=instance)
		#Check data.
		if form.has_changed():
			if form.is_valid():
				#Save data.
				epp = form.save()
				#We check if the epp was edited and notify to the user
				if epp:
					messages.success(request, 'Equipo de Protección personal editado correctamente')
					return redirect('epp_list')
				else:
					messages.error(request, 'Ocurrió un error y el E.P.P. no fué editado.')
			else:
				messages.error(request, 'Ocurrió un error, el formulario no es válido.')
		else:
			messages.success(request, 'Ningún cambio registrado.')
			return redirect('epp_list')
	#we pass the form with the instance to the view.
	form = EppEditForm(instance=instance)		
	return render(request, 'epps/epp_edit.html', 
					{ 'form' : form, 'title' : 'Editar E.P.P.'})

#This view is to delete a epp. Only admin users
#can access it
def epp_delete(request, pk):
	#We check if the logged user is admin
	if request.user.is_authenticated and request.user.is_superuser:
		#We take from url the pk of the epp.
		epp = EPP.objects.get(pk=pk)
		#We delete the epp and check
		if epp.delete():
			messages.success(request, 'Equipo de Protección Personal eliminado correctamente')
			return redirect('epp_list')
		else:
			messages.error(request, 'Ocurrió un error y el E.P.P. no fué eliminado.')
	else:
		return redirect('document')		

#### End of Views from EPP class	

#### Views from Hazard class
#This view is to create a new hazard. Only admin users
#can access it
def hazard_new(request):
	#we define the form
	form = HazardForm()
	if request.method == 'POST':
		form = HazardForm(request.POST)
		#Check data.
		if form.is_valid():
			#Save data.
			hazard = form.save()
			#We check if the hazard was created and notify to the user.
			if hazard:
				messages.success(request, 'Peligro o Riesgo creado correctamente')
				return redirect('hazard_new')
			else:
				messages.error(request, 'Ocurrió un error y el Peligro o Riesgo no fué creado.')
	
	return render(request, 'hazards/hazard_new.html', 
					{ 'form' : form, 'title' : 'Nuevo Peligro o Riesgo'})	

#This view shows in a list all hazards to edit and delete them.
#Is a view based in class.
class HazardListView(ListView):
	template_name = 'hazards/hazard_list.html'
	queryset = Hazard.objects.all().order_by('name')
	#We rewrite the context data to send to the view.
	def get_context_data(self, **kwargs):
		context = super().get_context_data(**kwargs)
		context['message'] = 'Listado de Peligros y Riesgos'
		context['title'] = 'Listado de Peligros y Riesgos'
		context['hazards'] = context['hazard_list']
		
		return context

#This view is to edit a hazard. Only admin users
#can access it
def hazard_edit(request, pk):
	#We take from url the pk of the hazard and 
	#generate an instance.
	instance = Hazard.objects.get(pk=pk)
	#We define the form
	form = HazardForm()
	if request.method == 'POST':
		form = HazardForm(request.POST, instance=instance)
		#Check data.
		if form.has_changed():
			if form.is_valid():
				#Save data.
				hazard = form.save()
				#We check if the hazard was edited and notify to the user
				if hazard:
					messages.success(request, 'Peligro o Riesgo editado correctamente')
					return redirect('hazard_list')
				else:
					messages.error(request, 'Ocurrió un error y el Peligro o Riesgo no fué editado.')
			else:
				messages.error(request, 'Ocurrió un error el formulario no es válido.')
		else:
			messages.success(request, 'Ningún cambio que registrar.')
			return redirect('hazard_list')

	form = HazardForm(instance=instance)		
	return render(request, 'hazards/hazard_edit.html', 
					{ 'form' : form , 'title' : 'Editar E.P.P.'})

#This view is to delete a hazard. Only admin users
#can access it
def hazard_delete(request, pk):
	#We check if the logged user is admin
	if request.user.is_authenticated and request.user.is_superuser:
		#We take from url the pk of the hazard.
		hazard = Hazard.objects.get(pk=pk)
		#We delete the hazard and check
		if hazard.delete():
			messages.success(request, 'Peligro o Riesgo eliminado correctamente')
			return redirect('hazard_list')
		else:
			messages.error(request, 'Ocurrió un error y el Peligro o Riesgo no fué eliminado.')
	else:
		return redirect('document')		

#### End of Views from Hazard class

#### Views from Precaution class
#This view is to create a new precaution. Only admin users
#can access it
def precaution_new(request):
	#We define the form
	form = PrecautionForm()
	if request.method == 'POST':
		form = PrecautionForm(request.POST)
		#Check data.
		if form.is_valid():
			#Save data.
			precaution = form.save()
			#We check if the precaution was created and notify to the user
			if precaution:
				messages.success(request, 'Medida Preventiva creada correctamente')
				return redirect('precaution_new')
			else:
				messages.error(request, 'Ocurrió un error y la Medida Preventiva no fué creada.')
	
	return render(request, 'precautions/precaution_new.html', 
					{ 'form' : form , 'title' : 'Nueva Medida Preventiva'})	

#This view shows in a list all precautions to edit and delete them.
#Is a view based in class.
class PrecautionListView(ListView):
	template_name = 'precautions/precaution_list.html'
	queryset = Precaution.objects.all().order_by('name')
	#We rewrite the context data to send to the view.
	def get_context_data(self, **kwargs):
		context = super().get_context_data(**kwargs)
		context['message'] = 'Listado de Medidas Preventivas'
		context['title'] = 'Listado de Medidas Preventivas'
		context['precautions'] = context['precaution_list']
		
		return context	

#This view is to edit a precaution. Only admin users
#can access it
def precaution_edit(request, pk):
	#We take from url the pk of the precaution and 
	#generate an instance.
	instance = Precaution.objects.get(pk=pk)
	#We define the form.
	form = PrecautionForm()
	if request.method == 'POST':
		form = PrecautionForm(request.POST, instance=instance)
		#Check data.
		if form.has_changed():
			if form.is_valid():
				#Save data
				precaution = form.save()
				#We check if the precaution was edited and notify to the user
				if precaution:
					messages.success(request, 'Medida Preventiva editada correctamente')
					return redirect('precaution_list')
				else:
					messages.error(request, 'Ocurrió un error y la Medida Preventiva no fué editada.')
			else:
				messages.error(request, 'Ocurrió un error el formulario no es válido.')
		else:
			messages.success(request, 'Ningún cambio que registrar.')
			return redirect('precaution_list')		
	#we pass the form with the instance to the view.
	form = PrecautionForm(instance=instance)		
	return render(request, 'precautions/precaution_edit.html', 
					{ 'form' : form , 'title' : 'Editar Medida Preventiva' })

#This view is to delete a precaution. Only admin users
#can access it
def precaution_delete(request, pk):
	#We check if the logged user is admin
	if request.user.is_authenticated and request.user.is_superuser:
		#We take from url the pk of the precaution.
		precaution = Precaution.objects.get(pk=pk)
		#We delete the precaution and check
		if precaution.delete():
			messages.success(request, 'Medida Preventiva eliminada correctamente')
			return redirect('precaution_list')
		else:
			messages.error(request, 'Ocurrió un error y la Medida Preventiva no fué eliminada.')
	else:
		return redirect('document')		


#### End of Views from Precaution class

#### Views from Tool class	
#This view is to create a new tool. Only admin users
#can access it
def tool_new(request):
	#We define the form.
	form = ToolForm()
	if request.method == 'POST':
		form = ToolForm(request.POST, request.FILES)
		#Check data.
		if form.is_valid():
			#Save data.
			tool = form.save()
			#We check if the tool was created and notify to the user
			if tool:
				messages.success(request, 'Herramienta creada correctamente')
				return redirect('tool_new')
			else:
				messages.error(request, 'Ocurrió un error y la Herramienta no fué creada.')
	
	return render(request, 'tools/tool_new.html', 
					{ 'form' : form , 'title' : 'Nueva Herramienta' })	

#This view shows in a list all tools to edit and delete them.
#Is a view based in class.
class ToolListView(ListView):
	template_name = 'tools/tool_list.html'
	queryset = Tool.objects.all().order_by('name')
	#We rewrite the context data to send to the view.
	def get_context_data(self, **kwargs):
		context = super().get_context_data(**kwargs)
		context['message'] = 'Listado de Herramientas'
		context['title'] = 'Listado de Herramientas'
		context['tools'] = context['tool_list']
		
		return context

#This view is to edit a tool. Only admin users
#can access it
def tool_edit(request, pk):
	#We take from url the pk of the tool and 
	#generate an instance.
	instance = Tool.objects.get(pk=pk)
	#We define the form.
	form = ToolEditForm()
	if request.method == 'POST':
		form = ToolEditForm(request.POST, request.FILES, instance=instance)
		#Check data.
		if form.has_changed():
			if form.is_valid():
				#Save data.
				tool = form.save()
				#We check if the tool was edited and notify to the user
				if tool:
					messages.success(request, 'Herramienta editada correctamente')
					return redirect('tool_list')
				else:
					messages.error(request, 'Ocurrió un error y la Herramienta no fué editada.')
			else:
				messages.error(request, 'Ocurrió un error el formulario no es válido.')
		else:
			messages.success(request, 'Ningún cambio que registrar.')
			return redirect('tool_list')
	#we pass the form with the instance to the view.		
	form = ToolEditForm(instance=instance)		
	return render(request, 'tools/tool_edit.html', 
						{ 'form' : form , 'title' : 'Editar Herramienta' })

#This view is to delete a tool. Only admin users
#can access it
def tool_delete(request, pk):
	#We check if the logged user is admin
	if request.user.is_authenticated and request.user.is_superuser:
		#We take from url the pk of the tool.
		tool = Tool.objects.get(pk=pk)
		#We delete the tool and check
		if tool.delete():
			messages.success(request, 'Herramienta eliminada correctamente')
			return redirect('tool_list')
		else:
			messages.error(request, 'Ocurrió un error y la Herramienta no fué eliminada.')
	else:
		return redirect('document')		


#### End of Views from Tool class

def document_export(request, pk):
	#We check if the logged user is admin
	if request.user.is_authenticated:
		#Get data form DB with pk form url
		document = APR.objects.get(pk=pk)
		documentLine = APRLine.objects.filter(apr_id=pk).order_by('id')

		#Borders and styles
		thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),
							top=Side(style='thin'),bottom=Side(style='thin'))
		medium_border = Border(left=Side(style='medium'),right=Side(style='medium'),
							top=Side(style='medium'),bottom=Side(style='medium'))
		greyFill = PatternFill(start_color='C0C0C0',
                   end_color='C0C0C0',
                   fill_type='solid')

		#We create the Work Book
		wb = Workbook()
		#We define actual sheet like the active sheet
		ws = wb.active
		#Sheet title
		ws.title = "APR - Frente"
		#Freeze fields of header
		#ws.freeze_panes = 'A13'

		####Header####It's too much, don´t you think?
		ws.merge_cells('B2:B5')
		ws['B2'] = 'Your Logo Here'
		ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
		ws['B2'].font = Font(size=16)
		ws['B2'].border = medium_border
		ws['B3'].border = medium_border
		ws['B4'].border = medium_border
		ws['B5'].border = medium_border

		ws.merge_cells('C2:E5')
		ws['C2'] = 'APR (Análisis Preliminar de Riesgos)'
		ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
		ws['C2'].font = Font(bold=True, size=20)
		ws['C2'].border = medium_border
		ws['D2'].border = medium_border
		ws['E2'].border = medium_border

		ws['F2'] = 'URUGUAY - APR - {}'.format(document.documentnumber)
		ws['F2'].alignment = Alignment(horizontal='left', vertical='center')
		ws['F2'].border = medium_border
	
		ws['F3'] = 'Revisión: A'
		ws['F3'].alignment = Alignment(horizontal='left', vertical='center')
		ws['F3'].border = medium_border

		ws['F4'] = 'Fecha de Revisión - 27-12-2017          Hoja 1 de 2'
		ws['F4'].alignment = Alignment(horizontal='left', vertical='center')
		ws['F4'].border = medium_border

		date = str(document.created_at)
		part = date.split(" ")[0].split("-")
		result = "/".join(reversed(part))
		ws['F5'] = 'Fecha: {}'.format(result)
		ws['F5'].alignment = Alignment(horizontal='left', vertical='center')
		ws['F5'].border = medium_border
		ws['F5'].font = Font(bold=True, size=12)
		ws['F5'].fill = greyFill

		ws.merge_cells('B6:F6')
		ws['B6'].border = medium_border
		ws['C6'].border = medium_border
		ws['D6'].border = medium_border
		ws['E6'].border = medium_border
		ws['F6'].border = medium_border

		ws.merge_cells('B7:E7')
		ws['B7'] = 'Área: '
		ws['B7'].alignment = Alignment(horizontal='left', vertical='center')
		ws['B7'].border = medium_border
		ws['B7'].font = Font(bold=True, size=12)
		ws['C7'].border = medium_border
		ws['D7'].border = medium_border
		ws['E7'].border = medium_border
		ws['B7'].fill = greyFill


		ws['F7'] = 'Equipo: '
		ws['F7'].alignment = Alignment(horizontal='left', vertical='center')
		ws['F7'].border = medium_border
		ws['F7'].font = Font(bold=True, size=12)
		ws['F7'].fill = greyFill

		ws.merge_cells('B8:E8')
		ws['B8'] = 'Servicios a ejecutar:  {}'.format(document.title)
		ws['B8'].alignment = Alignment(horizontal='left', vertical='center')
		ws['B8'].border = medium_border
		ws['B8'].font = Font(bold=True, size=12)

		ws['F8'] = 'Severidad:       A |_____|        B |_____|        C |_____| '
		ws['F8'].alignment = Alignment(horizontal='center', vertical='center')
		ws['F8'].border = medium_border
		ws['F8'].font = Font(bold=True, size=12)
		ws['F8'].fill = greyFill

		ws['F9'] = 'Estado:  {}'.format(document.status)
		ws['F9'].alignment = Alignment(horizontal='left', vertical='center')
		ws['F9'].border = medium_border
		ws['F9'].font = Font(bold=True, size=12)
		ws['F9'].fill = greyFill

		ws['B9'] = 'Creado por: '
		ws['B9'].alignment = Alignment(horizontal='center', vertical='center')
		ws['B9'].border = medium_border
		ws['B9'].font = Font(bold=True, size=12)
		ws['B9'].fill = greyFill

		ws.merge_cells('C9:E9')
		ws['C9'] = ' {} '.format(document.user)
		ws['C9'].alignment = Alignment(horizontal='center', vertical='center')
		ws['C9'].border = medium_border
		ws['D9'].border = medium_border
		ws['E9'].border = medium_border
	
		ws['F9'].border = medium_border

		ws.merge_cells('B10:B11')
		ws['B10'] = 'Comentarios: '
		ws['B10'].alignment = Alignment(horizontal='center', vertical='center')
		ws['B10'].border = medium_border
		ws['B10'].font = Font(bold=True, size=12)
		ws['B10'].fill = greyFill

		ws.merge_cells('C10:F11')
		ws['C10'] = document.comments
		ws['C10'].alignment = Alignment(horizontal='left', vertical='center')
		ws['C10'].border = medium_border
		ws['D10'].border = medium_border
		ws['E10'].border = medium_border
		ws['F10'].border = medium_border

		ws['B11'].border = medium_border
		ws['C11'].border = medium_border
		ws['D11'].border = medium_border
		ws['E11'].border = medium_border
		ws['F11'].border = medium_border

		ws['B12'] = 'ACTIVIDADES \n (con sus respectivas etapas, \n detallando COMO serán realizadas)'
		ws['B12'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
		ws['B12'].font = Font(bold=True)
		ws['B12'].border = medium_border
		ws['B12'].fill = greyFill

		ws['C12'] = 'E.P.P Especiales'
		ws['C12'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
		ws['C12'].font = Font(bold=True)
		ws['C12'].border = medium_border
		ws['C12'].fill = greyFill

		ws['D12'] = 'HERRAMIENTAS/EQUIPOS.'
		ws['D12'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
		ws['D12'].font = Font(bold=True)
		ws['D12'].border = medium_border
		ws['D12'].fill = greyFill

		ws['E12'] = 'RIESGOS POTENCIALES \n (Que podría salir mal)'
		ws['E12'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
		ws['E12'].font = Font(bold=True)
		ws['E12'].border = medium_border
		ws['E12'].fill = greyFill

		ws['F12'] = 'MEDIDAS PREVENTIVAS / RECOMENDACIONES DE SEGURIDAD \n (Evitar los accidentes o minimizar los \n daños, en caso que ocurra)'
		ws['F12'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
		ws['F12'].font = Font(bold=True)
		ws['F12'].border = medium_border
		ws['F12'].fill = greyFill

		####End Header#### Finally!!!

		####Body####
		num = 13
		for line in documentLine:
			ws.cell(row=num,column=2).value = line.activities
			ws.cell(row=num,column=2).border = thin_border
			ws.cell(row=num,column=2).alignment = Alignment(wrap_text=True, vertical='center',
															 horizontal='left')
			epp = ''
			for e in line.epps.all():
				epp = epp + '* ' + e.name + '\n'

			ws.cell(row=num,column=3).value = epp
			ws.cell(row=num,column=3).border = thin_border
			ws.cell(row=num,column=3).alignment = Alignment(wrap_text=True, vertical='center',
															 horizontal='left')
			tool = ''
			for t in line.tools.all():
				tool = tool + '* ' + t.name + '\n'
			ws.cell(row=num,column=4).value = tool
			ws.cell(row=num,column=4).border = thin_border
			ws.cell(row=num,column=4).alignment = Alignment(wrap_text=True, vertical='center',
															 horizontal='left')
			hazard = ''
			for h in line.hazards.all():
				hazard = hazard + '* ' + h.name + '\n'
			ws.cell(row=num,column=5).value = hazard
			ws.cell(row=num,column=5).border = thin_border
			ws.cell(row=num,column=5).alignment = Alignment(wrap_text=True, vertical='center',
															 horizontal='left')
			precaution = ''
			for p in line.precautions.all():
				precaution = precaution + '* ' + p.name + '\n'
			ws.cell(row=num,column=6).value = precaution
			ws.cell(row=num,column=6).border = thin_border
			ws.cell(row=num,column=6).alignment = Alignment(wrap_text=True, vertical='center',
															 horizontal='left')
			ws.row_dimensions[num].height = 100.0
			num = num + 1



		#We set the name of the file
		file_name = 'APR-{}xlsx'.format(str(document.documentnumber) + '.')
		#We define the type of response
		response = HttpResponse(content_type="application/ms-excel")
		data = "attachment; filename={0}".format(file_name)
		response["Content-Disposition"] = data
		#Dimensions of columns and rows
		ws.column_dimensions["A"].width = 1.0
		ws.column_dimensions["B"].width = 50.0
		ws.column_dimensions["C"].width = 25.0
		ws.column_dimensions["D"].width = 25.0
		ws.column_dimensions["E"].width = 25.0
		ws.column_dimensions["F"].width = 60.0
		ws.row_dimensions[1].height = 10.0
		ws.row_dimensions[6].height = 7.0
		counter = 7
		while counter < 12:
			ws.row_dimensions[counter].height = 25.0
			counter = counter + 1

		ws.row_dimensions[12].height = 47.0
	
		wb.save(response)
		#Return the excel file.
		return response

	else:
		return redirect('document')